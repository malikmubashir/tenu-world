#!/usr/bin/env python3
"""
Governance Sync — Paperclip DB → Excel Tracker
Reads issue states, agent states, and run history from Paperclip's embedded Postgres.
Updates the Tenu-Reste-a-Faire.xlsx workbook with current status.
Appends new Paperclip issues not yet in the tracker.
Generates a governance summary for Telegram digest.

Usage:
  python3 governance-sync.py              # sync once
  python3 governance-sync.py --report     # sync + send Telegram summary
  python3 governance-sync.py --loop       # continuous sync every 30 min
"""

import json
import os
import subprocess
import sys
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Configuration ──
TRACKER_PATH = Path.home() / "Documents" / "Claude" / "Projects" / "Tenu.World" / "Tenu-Reste-a-Faire.xlsx"
SYNC_LOG = Path.home() / ".paperclip" / "governance-sync.log"
ENV_FILE = Path.home() / ".paperclip-monitor.env"

PG_HOST = "localhost"
PG_PORT = "54329"
PG_USER = "paperclip"
PG_DB = "paperclip"
PG_PASS = "paperclip"
PSQL_BIN = "/opt/homebrew/Cellar/libpq/18.3/bin/psql"

SYNC_INTERVAL_SEC = 1800  # 30 minutes

# Status mapping: Paperclip → Excel
STATUS_MAP = {
    "in_progress": "In Progress",
    "todo": "To Do",
    "done": "Done",
    "backlog": "To Do",
    "blocked": "Blocked",
    "cancelled": "Done",
}

# Style definitions
FILLS = {
    "Done": PatternFill("solid", fgColor="EAFAF1"),
    "In Progress": PatternFill("solid", fgColor="FEF9E7"),
    "Blocked": PatternFill("solid", fgColor="FDEDEC"),
    "To Do": PatternFill("solid", fgColor="EBF5FB"),
    "Review": PatternFill("solid", fgColor="FDF2E9"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    try:
        SYNC_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(SYNC_LOG, "a") as f:
            f.write(line + "\n")
    except Exception:
        pass


# ── Postgres query helper ──
def pg_query(sql):
    env = os.environ.copy()
    env["PGPASSWORD"] = PG_PASS
    cmd = [
        PSQL_BIN, "-h", PG_HOST, "-p", PG_PORT, "-U", PG_USER, "-d", PG_DB,
        "-t", "-A", "-F", "\t", "-c", sql,
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15, env=env)
        if result.returncode != 0:
            log(f"PSQL error: {result.stderr.strip()}")
            return []
        rows = []
        for line in result.stdout.strip().splitlines():
            if line.strip():
                rows.append(line.split("\t"))
        return rows
    except subprocess.TimeoutExpired:
        log("PSQL query timed out")
        return []
    except FileNotFoundError:
        log(f"psql not found at {PSQL_BIN}")
        return []


# ── Telegram sender ──
def load_telegram_config():
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, val = line.split("=", 1)
                os.environ.setdefault(key.strip(), val.strip())
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    return token, chat_id


def send_telegram(token, chat_id, message):
    if not token or not chat_id:
        return False
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = json.dumps({"chat_id": chat_id, "text": message}).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status == 200
    except Exception as e:
        log(f"Telegram send failed: {e}")
        return False


# ── Data collection from Paperclip ──
def collect_paperclip_state():
    """Collect all relevant state from Paperclip DB."""
    state = {}

    # Agent statuses
    state["agents"] = []
    for row in pg_query("""
        SELECT name, status, adapter_type, adapter_config->>'model' as model,
               CASE WHEN paused_at IS NOT NULL THEN 'paused' ELSE 'active' END as pause_state,
               budget_monthly_cents, spent_monthly_cents
        FROM agents ORDER BY name;
    """):
        if len(row) >= 7:
            state["agents"].append({
                "name": row[0], "status": row[1], "adapter": row[2],
                "model": row[3], "pause_state": row[4],
                "budget_cents": int(row[5] or 0), "spent_cents": int(row[6] or 0),
            })

    # All issues with assignees
    state["issues"] = []
    for row in pg_query("""
        SELECT i.id, i.title, i.status, COALESCE(a.name, 'unassigned') as assignee,
               i.created_at::text, i.updated_at::text
        FROM issues i
        LEFT JOIN agents a ON i.assignee_agent_id = a.id
        ORDER BY i.created_at;
    """):
        if len(row) >= 6:
            state["issues"].append({
                "id": row[0], "title": row[1], "status": row[2],
                "assignee": row[3], "created": row[4], "updated": row[5],
            })

    # Recent runs (last 24h)
    state["recent_runs"] = []
    for row in pg_query("""
        SELECT a.name, hr.status, hr.error,
               EXTRACT(EPOCH FROM (COALESCE(hr.finished_at, NOW()) - hr.started_at))::int as elapsed,
               hr.started_at::text, hr.finished_at::text
        FROM heartbeat_runs hr
        JOIN agents a ON hr.agent_id = a.id
        WHERE hr.started_at > NOW() - INTERVAL '24 hours'
        ORDER BY hr.started_at DESC;
    """):
        if len(row) >= 6:
            state["recent_runs"].append({
                "agent": row[0], "status": row[1], "error": row[2],
                "elapsed": int(row[3] or 0), "started": row[4], "finished": row[5],
            })

    # Run stats (all time)
    state["run_stats"] = {}
    for row in pg_query("""
        SELECT status, COUNT(*) FROM heartbeat_runs GROUP BY status ORDER BY status;
    """):
        if len(row) >= 2:
            state["run_stats"][row[0]] = int(row[1])

    return state


# ── Excel sync ──
def sync_excel(state):
    """Update the Excel tracker with current Paperclip state."""
    if not TRACKER_PATH.exists():
        log(f"Tracker not found at {TRACKER_PATH}")
        return False

    wb = openpyxl.load_workbook(TRACKER_PATH)

    # ── Update Reste-à-Faire sheet ──
    if "Reste-à-Faire" in wb.sheetnames:
        ws = wb["Reste-à-Faire"]
        update_raf_sheet(ws, state)

    # ── Update or create Paperclip Status sheet ──
    if "Paperclip Status" in wb.sheetnames:
        wb.remove(wb["Paperclip Status"])
    ws_status = wb.create_sheet("Paperclip Status")
    ws_status.sheet_properties.tabColor = "8E44AD"
    build_status_sheet(ws_status, state)

    # ── Update or create Run History sheet ──
    if "Run History" in wb.sheetnames:
        wb.remove(wb["Run History"])
    ws_runs = wb.create_sheet("Run History")
    ws_runs.sheet_properties.tabColor = "2980B9"
    build_run_history_sheet(ws_runs, state)

    wb.save(TRACKER_PATH)
    log(f"Excel tracker updated: {TRACKER_PATH}")
    return True


def update_raf_sheet(ws, state):
    """Cross-reference Paperclip issues with RAF tasks and update statuses."""
    # Build lookup: Paperclip issue title → status
    issue_lookup = {}
    for issue in state["issues"]:
        # Normalize title for fuzzy matching
        key = issue["title"].strip().lower()[:60]
        issue_lookup[key] = issue

    updated_count = 0
    # Scan RAF rows (skip header)
    for row_num in range(2, ws.max_row + 1):
        task_cell = ws.cell(row=row_num, column=3)  # Column C = Task
        status_cell = ws.cell(row=row_num, column=5)  # Column E = Status
        notes_cell = ws.cell(row=row_num, column=10)  # Column J = Notes

        if not task_cell.value:
            continue

        task_title = str(task_cell.value).strip().lower()[:60]
        current_status = str(status_cell.value or "").strip()

        # Try to match with Paperclip issue
        matched_issue = None
        for key, issue in issue_lookup.items():
            # Fuzzy match: check if significant words overlap
            task_words = set(task_title.split())
            issue_words = set(key.split())
            overlap = task_words & issue_words
            if len(overlap) >= 3 or key in task_title or task_title in key:
                matched_issue = issue
                break

        if matched_issue:
            new_status = STATUS_MAP.get(matched_issue["status"], current_status)
            if new_status != current_status and current_status != "Done":
                status_cell.value = new_status
                # Apply fill
                fill = FILLS.get(new_status)
                if fill:
                    for c in range(1, 11):
                        ws.cell(row=row_num, column=c).fill = fill

                # Add sync note
                sync_note = f"[Sync {datetime.now().strftime('%m/%d %H:%M')}] Paperclip: {matched_issue['status']}"
                existing_notes = str(notes_cell.value or "")
                if "Sync" not in existing_notes:
                    notes_cell.value = sync_note
                else:
                    notes_cell.value = sync_note

                updated_count += 1

    # Append Paperclip issues not in RAF
    existing_titles = set()
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=3).value
        if val:
            existing_titles.add(str(val).strip().lower()[:60])

    new_count = 0
    for issue in state["issues"]:
        if issue["status"] == "done":
            continue
        title_key = issue["title"].strip().lower()[:60]
        # Check if already tracked
        found = False
        for existing in existing_titles:
            words_a = set(title_key.split())
            words_b = set(existing.split())
            if len(words_a & words_b) >= 3:
                found = True
                break
        if not found:
            next_row = ws.max_row + 1
            # Find next ID
            last_id = "T-099"
            for r in range(ws.max_row, 1, -1):
                v = ws.cell(row=r, column=1).value
                if v and str(v).startswith("T-"):
                    last_id = str(v)
                    break
            id_num = int(last_id.split("-")[1]) + 1 + new_count
            new_id = f"T-{id_num:03d}"

            excel_status = STATUS_MAP.get(issue["status"], "To Do")
            ws.cell(row=next_row, column=1, value=new_id)
            ws.cell(row=next_row, column=2, value="Paperclip")
            ws.cell(row=next_row, column=3, value=issue["title"])
            ws.cell(row=next_row, column=4, value=issue["assignee"])
            ws.cell(row=next_row, column=5, value=excel_status)
            ws.cell(row=next_row, column=6, value="P1")
            ws.cell(row=next_row, column=10, value=f"[Auto-added from Paperclip {datetime.now().strftime('%m/%d')}]")

            fill = FILLS.get(excel_status)
            for c in range(1, 11):
                cell = ws.cell(row=next_row, column=c)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill

            new_count += 1

    log(f"RAF: {updated_count} statuses updated, {new_count} new issues added")


def build_status_sheet(ws, state):
    """Build a live Paperclip status snapshot sheet."""
    ws["A1"] = f"PAPERCLIP STATUS — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B2A4A")

    # Agent status table
    headers = ["Agent", "Status", "Adapter", "Model", "State", "Budget", "Spent"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1B2A4A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, agent in enumerate(state["agents"]):
        r = 4 + i
        ws.cell(row=r, column=1, value=agent["name"])
        ws.cell(row=r, column=2, value=agent["status"])
        ws.cell(row=r, column=3, value=agent["adapter"])
        ws.cell(row=r, column=4, value=agent["model"] or "n/a")
        ws.cell(row=r, column=5, value=agent["pause_state"].upper())
        ws.cell(row=r, column=6, value=f"${agent['budget_cents']/100:.0f}")
        ws.cell(row=r, column=7, value=f"${agent['spent_cents']/100:.2f}")

        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.border = THIN_BORDER
            if agent["pause_state"] == "paused":
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
                cell.font = Font(name="Arial", size=10, color="999999")

    # Issue summary
    issue_row = 4 + len(state["agents"]) + 2
    ws.cell(row=issue_row, column=1, value="ISSUE SUMMARY")
    ws.cell(row=issue_row, column=1).font = Font(name="Arial", bold=True, size=12, color="1B2A4A")

    status_counts = {}
    for issue in state["issues"]:
        s = issue["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    sr = issue_row + 1
    for status, count in sorted(status_counts.items()):
        ws.cell(row=sr, column=1, value=status)
        ws.cell(row=sr, column=2, value=count)
        sr += 1

    # Run stats
    sr += 1
    ws.cell(row=sr, column=1, value="RUN STATISTICS (ALL TIME)")
    ws.cell(row=sr, column=1).font = Font(name="Arial", bold=True, size=12, color="1B2A4A")
    sr += 1
    for status, count in sorted(state["run_stats"].items()):
        ws.cell(row=sr, column=1, value=status)
        ws.cell(row=sr, column=2, value=count)
        sr += 1

    for w, c in [(25, 1), (12, 2), (15, 3), (15, 4), (10, 5), (10, 6), (10, 7)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w


def build_run_history_sheet(ws, state):
    """Build a run history sheet from recent runs."""
    ws["A1"] = f"RUN HISTORY (Last 24h) — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1B2A4A")

    headers = ["Agent", "Status", "Duration", "Error", "Started", "Finished"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1B2A4A")
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for i, run in enumerate(state["recent_runs"]):
        r = 4 + i
        elapsed_min = run["elapsed"] // 60
        ws.cell(row=r, column=1, value=run["agent"])
        ws.cell(row=r, column=2, value=run["status"])
        ws.cell(row=r, column=3, value=f"{elapsed_min}m")
        ws.cell(row=r, column=4, value=run["error"] or "")
        ws.cell(row=r, column=5, value=run["started"][:19] if run["started"] else "")
        ws.cell(row=r, column=6, value=run["finished"][:19] if run["finished"] else "")

        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if run["status"] == "succeeded":
                cell.fill = PatternFill("solid", fgColor="EAFAF1")
            elif run["status"] in ("failed", "timed_out"):
                cell.fill = PatternFill("solid", fgColor="FDEDEC")

    for w, c in [(25, 1), (12, 2), (10, 3), (50, 4), (20, 5), (20, 6)]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w


# ── Governance report ──
def build_governance_report(state):
    """Build a Telegram governance summary."""
    now_str = datetime.now().strftime("%H:%M")
    lines = [f"TENU GOVERNANCE REPORT  {now_str}", ""]

    # Agent states
    active = [a for a in state["agents"] if a["pause_state"] == "active"]
    paused = [a for a in state["agents"] if a["pause_state"] == "paused"]
    lines.append(f"Agents: {len(active)} active, {len(paused)} paused")
    for a in active:
        budget_str = ""
        if a["budget_cents"] > 0:
            pct = round(a["spent_cents"] / a["budget_cents"] * 100) if a["budget_cents"] > 0 else 0
            budget_str = f"  ${a['spent_cents']/100:.2f}/${a['budget_cents']/100:.0f} ({pct}%)"
        lines.append(f"  {a['name']} [{a['model'] or a['adapter']}]{budget_str}")

    # Issue breakdown
    lines.append("")
    status_counts = {}
    for issue in state["issues"]:
        s = issue["status"]
        status_counts[s] = status_counts.get(s, 0) + 1
    total = sum(status_counts.values())
    parts = [f"{total} total"]
    for s in ["in_progress", "todo", "done", "blocked", "backlog"]:
        if s in status_counts:
            parts.append(f"{status_counts[s]} {s}")
    lines.append("Issues: " + ", ".join(parts))

    # Reste-à-faire count (non-done)
    raf_count = sum(1 for i in state["issues"] if i["status"] not in ("done", "cancelled"))
    lines.append(f"Reste-a-faire: {raf_count} tasks pending")

    # Recent runs (24h)
    runs_24h = state["recent_runs"]
    if runs_24h:
        succeeded = sum(1 for r in runs_24h if r["status"] == "succeeded")
        failed = sum(1 for r in runs_24h if r["status"] in ("failed", "timed_out"))
        lines.append("")
        lines.append(f"Runs (24h): {len(runs_24h)} total, {succeeded} ok, {failed} failed")

    # Blockers
    blocked = [i for i in state["issues"] if i["status"] == "blocked"]
    if blocked:
        lines.append("")
        lines.append("BLOCKED:")
        for b in blocked:
            lines.append(f"  {b['assignee']}: {b['title'][:50]}")

    lines.append("")
    lines.append(f"Tracker: Tenu-Reste-a-Faire.xlsx (auto-synced)")

    return "\n".join(lines)


# ── Main ──
def run_sync(send_report=False):
    log("Starting governance sync...")

    state = collect_paperclip_state()
    if not state["agents"]:
        log("No data from Paperclip DB. Is the server running?")
        return

    log(f"Collected: {len(state['agents'])} agents, {len(state['issues'])} issues, {len(state['recent_runs'])} recent runs")

    if sync_excel(state):
        log("Excel sync complete.")

    if send_report:
        token, chat_id = load_telegram_config()
        report = build_governance_report(state)
        if send_telegram(token, chat_id, report):
            log("Governance report sent to Telegram.")
        else:
            log("Telegram report skipped (no config or send failed).")
            print(report)

    log("Sync complete.")


if __name__ == "__main__":
    if "--loop" in sys.argv:
        print(f"Governance Sync running (every {SYNC_INTERVAL_SEC}s). Ctrl+C to stop.")
        while True:
            try:
                run_sync(send_report="--report" in sys.argv)
                time.sleep(SYNC_INTERVAL_SEC)
            except KeyboardInterrupt:
                print("\nStopped.")
                break
    else:
        run_sync(send_report="--report" in sys.argv)
