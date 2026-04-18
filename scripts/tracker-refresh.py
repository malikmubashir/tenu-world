#!/usr/bin/env python3
"""
tracker-refresh.py

Regenerates Tenu-Reste-a-Faire.xlsx from TASKS.md.

TASKS.md is the master. This xlsx is the human-readable mirror.
Run at the end of any session that mutated TASKS.md.

Usage:
    python3 scripts/tracker-refresh.py

Exit codes:
    0 success
    1 TASKS.md not found or unparseable
    2 openpyxl missing
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. pip install openpyxl", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parent.parent
TASKS_MD = REPO / "TASKS.md"
XLSX = REPO / "Tenu-Reste-a-Faire.xlsx"
DASHBOARD_HTML = REPO / "dashboard.html"
LAUNCH_DATE = date(2026, 5, 11)  # Fork B soft launch target

STATUS_MAP = {
    " ": "Open",
    "x": "Done",
    "X": "Done",
    ">": "In Progress",
    "-": "Dropped",
}

PRIORITY_MAP = {
    "0": "P0 Blocker",
    "1": "P1 Important",
    "2": "P2 Post-launch",
}

# line like: - [ ] MH: Title body (p:0, due 2026-04-18) — trailing note
TASK_LINE = re.compile(
    r"^-\s+\[([ xX>\-])\]\s+(.*)$"
)
OWNER_PREFIX = re.compile(
    r"^(MH|CC)(?:\s*\([^)]+\))?:\s*(.*)$"
)
META_BLOCK = re.compile(
    r"\(([^)]*p:[012][^)]*)\)"
)
PRIORITY_RE = re.compile(r"p:([012])")
DUE_RE = re.compile(r"due\s+(\d{4}-\d{2}-\d{2})")


@dataclass
class Task:
    line_no: int
    section: str
    status: str
    owner: str
    title: str
    priority: Optional[str]
    due: Optional[date]
    note: str

    def task_id(self, idx: int) -> str:
        return f"T-{idx:03d}"


def parse_tasks(text: str) -> tuple[list[Task], dict]:
    tasks: list[Task] = []
    meta: dict = {}
    section = "(unsectioned)"
    for line_no, raw in enumerate(text.splitlines(), start=1):
        line = raw.rstrip()
        if line.startswith("# "):
            meta.setdefault("title", line[2:].strip())
            continue
        if line.startswith("## "):
            section = line[3:].strip()
            continue
        m = TASK_LINE.match(line)
        if not m:
            continue
        state_char, body = m.group(1), m.group(2).strip()
        status = STATUS_MAP.get(state_char, "Unknown")
        owner_match = OWNER_PREFIX.match(body)
        if not owner_match:
            continue
        owner = owner_match.group(1)
        rest = owner_match.group(2).strip()

        # First: extract the (p:X, due YYYY-MM-DD) block from anywhere in the body.
        # The metadata parenthesis can sit before or after the em dash separator.
        priority = None
        due: Optional[date] = None
        meta_match = META_BLOCK.search(rest)
        if meta_match:
            inner = meta_match.group(1)
            p = PRIORITY_RE.search(inner)
            d = DUE_RE.search(inner)
            if p:
                priority = PRIORITY_MAP.get(p.group(1))
            if d:
                try:
                    due = datetime.strptime(d.group(1), "%Y-%m-%d").date()
                except ValueError:
                    due = None
            # Remove the meta block from the body so it doesn't leak into title/note.
            rest = (rest[: meta_match.start()] + rest[meta_match.end():]).strip()
            # Tidy stray whitespace/punctuation left where the block used to sit.
            rest = re.sub(r"\s{2,}", " ", rest).strip(" ,")

        # Then: split trailing note on " — " (em dash with spaces). First occurrence
        # separates title from status note. Em dash is reserved for notes per
        # TASKS.md convention, so later em dashes inside notes are fine.
        if " — " in rest:
            title, _, note = rest.partition(" — ")
        else:
            title, note = rest, ""
        title = title.strip().rstrip(", ")
        note = note.strip()

        tasks.append(Task(
            line_no=line_no,
            section=section,
            status=status,
            owner=owner,
            title=title,
            priority=priority,
            due=due,
            note=note,
        ))
    return tasks, meta


# ---------- xlsx rendering ----------

PALETTE = {
    "navy": "FF0B1F3A",
    "paper": "FFF4F1EA",
    "emerald": "FF059669",
    "ink": "FF1D1D1F",
    "mute": "FFE5E7EB",
    "red": "FFB91C1C",
    "amber": "FFB45309",
    "green": "FF047857",
    "gray": "FF6B7280",
}

HEADER_FILL = PatternFill("solid", fgColor=PALETTE["navy"])
HEADER_FONT = Font(bold=True, color="FFFFFFFF", name="Inter", size=11)
CELL_FONT = Font(name="Inter", size=10, color=PALETTE["ink"])
SECTION_FONT = Font(bold=True, name="Inter", size=11, color=PALETTE["navy"])
TITLE_FONT = Font(bold=True, name="Inter", size=14, color=PALETTE["navy"])
SUBTITLE_FONT = Font(italic=True, name="Inter", size=10, color=PALETTE["gray"])

STATUS_FONT = {
    "Open": Font(name="Inter", size=10, color=PALETTE["ink"]),
    "In Progress": Font(bold=True, name="Inter", size=10, color=PALETTE["amber"]),
    "Done": Font(name="Inter", size=10, color=PALETTE["green"]),
    "Dropped": Font(italic=True, name="Inter", size=10, color=PALETTE["gray"]),
}


def set_col_widths(ws, widths: list[tuple[int, int]]):
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


def render_summary(ws, tasks: list[Task], meta: dict):
    ws.title = "Priority Summary"
    ws["A1"] = "TENU.WORLD — TASK TRACKER"
    ws["A1"].font = TITLE_FONT
    today = date.today().isoformat()
    ws["A2"] = f"Regenerated {today} from TASKS.md. Master = TASKS.md. Do not edit xlsx directly."
    ws["A2"].font = SUBTITLE_FONT

    # counts
    open_tasks = [t for t in tasks if t.status in ("Open", "In Progress")]
    done_tasks = [t for t in tasks if t.status == "Done"]
    dropped = [t for t in tasks if t.status == "Dropped"]

    def count(predicate):
        return sum(1 for t in open_tasks if predicate(t))

    row = 4
    ws.cell(row=row, column=1, value="BY STATUS").font = SECTION_FONT
    row += 1
    for header in ["Status", "Count"]:
        c = ws.cell(row=row, column=["Status", "Count"].index(header) + 1, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    for label, n in [
        ("Open", count(lambda t: t.status == "Open")),
        ("In Progress", count(lambda t: t.status == "In Progress")),
        ("Done", len(done_tasks)),
        ("Dropped", len(dropped)),
    ]:
        ws.cell(row=row, column=1, value=label).font = CELL_FONT
        ws.cell(row=row, column=2, value=n).font = CELL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="BY PRIORITY (open only)").font = SECTION_FONT
    row += 1
    for header in ["Priority", "Count", "Note"]:
        c = ws.cell(row=row, column=["Priority", "Count", "Note"].index(header) + 1, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    notes = {
        "P0 Blocker": "Must clear before soft launch 11 May 2026",
        "P1 Important": "Nice to have at launch, not blocking",
        "P2 Post-launch": "After 11 May",
    }
    for label in ["P0 Blocker", "P1 Important", "P2 Post-launch"]:
        ws.cell(row=row, column=1, value=label).font = CELL_FONT
        ws.cell(row=row, column=2, value=count(lambda t, l=label: t.priority == l)).font = CELL_FONT
        ws.cell(row=row, column=3, value=notes[label]).font = CELL_FONT
        row += 1
    # priority None
    nless = count(lambda t: t.priority is None)
    if nless:
        ws.cell(row=row, column=1, value="No priority").font = CELL_FONT
        ws.cell(row=row, column=2, value=nless).font = CELL_FONT
        ws.cell(row=row, column=3, value="Tag with p:0, p:1 or p:2 in TASKS.md").font = CELL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="BY OWNER (open only)").font = SECTION_FONT
    row += 1
    for header in ["Owner", "Count"]:
        c = ws.cell(row=row, column=["Owner", "Count"].index(header) + 1, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    for label in ["MH", "CC"]:
        ws.cell(row=row, column=1, value=label).font = CELL_FONT
        ws.cell(row=row, column=2, value=count(lambda t, l=label: t.owner == l)).font = CELL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="NEXT 7 DAYS (open + due within 7 days)").font = SECTION_FONT
    row += 1
    for header in ["Due", "Owner", "Priority", "Title"]:
        c = ws.cell(row=row, column=["Due", "Owner", "Priority", "Title"].index(header) + 1, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    horizon = date.today().toordinal() + 7
    due_soon = [t for t in open_tasks if t.due and t.due.toordinal() <= horizon]
    due_soon.sort(key=lambda t: (t.due or date.max, t.priority or "ZZ"))
    if not due_soon:
        ws.cell(row=row, column=1, value="(nothing due in next 7 days)").font = CELL_FONT
        row += 1
    for t in due_soon:
        ws.cell(row=row, column=1, value=t.due.isoformat() if t.due else "").font = CELL_FONT
        ws.cell(row=row, column=2, value=t.owner).font = CELL_FONT
        ws.cell(row=row, column=3, value=t.priority or "").font = CELL_FONT
        ws.cell(row=row, column=4, value=t.title).font = CELL_FONT
        row += 1

    set_col_widths(ws, [(1, 22), (2, 14), (3, 18), (4, 90)])
    ws.freeze_panes = "A4"


def render_tasks_sheet(ws, tasks: list[Task], title: str, filter_fn):
    ws.title = title
    headers = ["ID", "Section", "Status", "Owner", "Priority", "Due", "Title", "Note"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    filtered = [t for t in tasks if filter_fn(t)]
    for idx, t in enumerate(filtered, start=1):
        r = idx + 1
        ws.cell(row=r, column=1, value=t.task_id(idx)).font = CELL_FONT
        ws.cell(row=r, column=2, value=t.section).font = CELL_FONT
        sc = ws.cell(row=r, column=3, value=t.status)
        sc.font = STATUS_FONT.get(t.status, CELL_FONT)
        ws.cell(row=r, column=4, value=t.owner).font = CELL_FONT
        ws.cell(row=r, column=5, value=t.priority or "").font = CELL_FONT
        due_val = t.due.isoformat() if t.due else ""
        ws.cell(row=r, column=6, value=due_val).font = CELL_FONT
        title_cell = ws.cell(row=r, column=7, value=t.title)
        title_cell.font = CELL_FONT
        title_cell.alignment = Alignment(wrap_text=True, vertical="top")
        note_cell = ws.cell(row=r, column=8, value=t.note)
        note_cell.font = CELL_FONT
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    set_col_widths(ws, [(1, 8), (2, 28), (3, 14), (4, 8), (5, 16), (6, 12), (7, 80), (8, 80)])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(filtered) + 1)}"


def render(tasks: list[Task], meta: dict, path: Path):
    wb = Workbook()
    render_summary(wb.active, tasks, meta)
    render_tasks_sheet(wb.create_sheet(), tasks, "Open",
                        lambda t: t.status in ("Open", "In Progress"))
    render_tasks_sheet(wb.create_sheet(), tasks, "Done",
                        lambda t: t.status == "Done")
    render_tasks_sheet(wb.create_sheet(), tasks, "Dropped",
                        lambda t: t.status == "Dropped")
    wb.save(path)


# ---------- html dashboard rendering ----------

import json


def task_to_dict(t: Task, idx: int) -> dict:
    return {
        "id": t.task_id(idx),
        "line_no": t.line_no,
        "section": t.section,
        "status": t.status,
        "owner": t.owner,
        "title": t.title,
        "priority": t.priority or "",
        "due": t.due.isoformat() if t.due else "",
        "note": t.note,
    }


def render_html(tasks: list[Task], meta: dict, path: Path) -> None:
    today = date.today()
    days_to_launch = (LAUNCH_DATE - today).days

    open_tasks = [t for t in tasks if t.status in ("Open", "In Progress")]
    done_tasks = [t for t in tasks if t.status == "Done"]
    in_progress = [t for t in tasks if t.status == "In Progress"]
    dropped = [t for t in tasks if t.status == "Dropped"]

    horizon = today.toordinal() + 7
    due_next_7 = [t for t in open_tasks if t.due and t.due.toordinal() <= horizon]
    overdue = [t for t in open_tasks if t.due and t.due.toordinal() < today.toordinal()]

    p0_open = [t for t in open_tasks if t.priority == "P0 Blocker"]
    p1_open = [t for t in open_tasks if t.priority == "P1 Important"]
    p2_open = [t for t in open_tasks if t.priority == "P2 Post-launch"]
    no_priority = [t for t in open_tasks if not t.priority]

    mh_open = [t for t in open_tasks if t.owner == "MH"]
    cc_open = [t for t in open_tasks if t.owner == "CC"]

    p0_no_due = [t for t in p0_open if not t.due]

    # Build task list with stable IDs for all tasks (not just open)
    all_tasks_json = [task_to_dict(t, i + 1) for i, t in enumerate(tasks)]

    summary = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "today": today.isoformat(),
        "launch": LAUNCH_DATE.isoformat(),
        "days_to_launch": days_to_launch,
        "counts": {
            "total": len(tasks),
            "open": len([t for t in tasks if t.status == "Open"]),
            "in_progress": len(in_progress),
            "done": len(done_tasks),
            "dropped": len(dropped),
            "p0_open": len(p0_open),
            "p1_open": len(p1_open),
            "p2_open": len(p2_open),
            "no_priority": len(no_priority),
            "mh_open": len(mh_open),
            "cc_open": len(cc_open),
            "due_next_7": len(due_next_7),
            "overdue": len(overdue),
            "p0_no_due": len(p0_no_due),
        },
    }

    payload = {
        "summary": summary,
        "tasks": all_tasks_json,
    }

    data_json = json.dumps(payload, ensure_ascii=False, indent=2)

    html = HTML_TEMPLATE.replace("__DATA_JSON__", data_json)
    html = html.replace("__GENERATED__", summary["generated"])
    html = html.replace("__DAYS_TO_LAUNCH__", str(days_to_launch))
    html = html.replace("__LAUNCH_DATE__", LAUNCH_DATE.isoformat())
    html = html.replace("__P0_REMAINING__", str(len(p0_open)))
    html = html.replace("__OPEN_TOTAL__", str(len(open_tasks)))
    html = html.replace("__IN_PROGRESS__", str(len(in_progress)))
    html = html.replace("__DONE_TOTAL__", str(len(done_tasks)))
    html = html.replace("__DUE_7__", str(len(due_next_7)))
    html = html.replace("__OVERDUE__", str(len(overdue)))
    html = html.replace("__P0_NO_DUE__", str(len(p0_no_due)))
    html = html.replace("__MH_OPEN__", str(len(mh_open)))
    html = html.replace("__CC_OPEN__", str(len(cc_open)))
    html = html.replace("__NO_PRIORITY__", str(len(no_priority)))

    path.write_text(html, encoding="utf-8")


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>tenu — task tracker</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  :root {
    --navy: #0B1F3A;
    --paper: #F4F1EA;
    --emerald: #059669;
    --emerald-dark: #047857;
    --ink: #1D1D1F;
    --mute: #E5E7EB;
    --mute-2: #F3F4F6;
    --red: #B91C1C;
    --amber: #B45309;
    --green: #047857;
    --gray: #6B7280;
    --gray-2: #9CA3AF;
  }
  * { box-sizing: border-box; }
  html, body {
    margin: 0;
    padding: 0;
    background: var(--paper);
    color: var(--ink);
    font-family: -apple-system, "SF Pro Text", "Inter", "Helvetica Neue", Arial, sans-serif;
    font-size: 14px;
    line-height: 1.5;
  }
  a { color: var(--emerald); text-decoration: none; }
  a:hover { text-decoration: underline; }

  .wrap { max-width: 1400px; margin: 0 auto; padding: 28px 32px 64px; }

  header.top {
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
    padding-bottom: 20px;
    border-bottom: 1px solid var(--mute);
    margin-bottom: 24px;
  }
  .brand {
    display: flex;
    align-items: baseline;
    gap: 12px;
  }
  .wordmark {
    font-family: "Inter Tight", "Inter", -apple-system, sans-serif;
    font-weight: 500;
    font-size: 28px;
    letter-spacing: -0.04em;
    color: var(--navy);
  }
  .subtitle {
    color: var(--gray);
    font-size: 13px;
  }
  .meta {
    text-align: right;
    font-size: 12px;
    color: var(--gray);
  }
  .meta .launch {
    color: var(--navy);
    font-weight: 600;
    font-size: 14px;
  }
  .meta .countdown-big {
    font-family: "Inter Tight", "Inter", sans-serif;
    font-size: 32px;
    font-weight: 600;
    color: var(--emerald);
    letter-spacing: -0.02em;
    line-height: 1;
  }
  .meta .countdown-big.urgent { color: var(--red); }

  .cards {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 16px;
    margin-bottom: 24px;
  }
  .card {
    background: #fff;
    border: 1px solid var(--mute);
    border-radius: 10px;
    padding: 18px 20px;
  }
  .card .label {
    text-transform: uppercase;
    font-size: 11px;
    letter-spacing: 0.08em;
    color: var(--gray);
    font-weight: 600;
  }
  .card .value {
    font-family: "Inter Tight", "Inter", sans-serif;
    font-size: 40px;
    font-weight: 600;
    color: var(--navy);
    letter-spacing: -0.02em;
    line-height: 1.1;
    margin-top: 4px;
  }
  .card .hint { color: var(--gray); font-size: 12px; margin-top: 4px; }
  .card.accent-emerald .value { color: var(--emerald); }
  .card.accent-amber .value { color: var(--amber); }
  .card.accent-red .value { color: var(--red); }

  .radar {
    background: var(--navy);
    color: #fff;
    border-radius: 10px;
    padding: 20px 22px;
    margin-bottom: 24px;
    display: grid;
    grid-template-columns: 1.5fr 1fr 1fr 1fr;
    gap: 24px;
    align-items: center;
  }
  .radar h3 {
    margin: 0 0 6px;
    font-family: "Inter Tight", "Inter", sans-serif;
    font-weight: 600;
    font-size: 18px;
    letter-spacing: -0.01em;
  }
  .radar .tagline { color: #C6CEDB; font-size: 12px; }
  .radar .stat .n {
    font-family: "Inter Tight", "Inter", sans-serif;
    font-size: 28px;
    font-weight: 600;
  }
  .radar .stat .n.warn { color: #F59E0B; }
  .radar .stat .n.bad { color: #FCA5A5; }
  .radar .stat .lbl { color: #C6CEDB; font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em; }

  .controls {
    display: flex;
    align-items: center;
    gap: 12px;
    flex-wrap: wrap;
    margin-bottom: 16px;
  }
  .pills {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
  }
  .pill {
    background: #fff;
    border: 1px solid var(--mute);
    color: var(--ink);
    padding: 6px 12px;
    border-radius: 999px;
    font-size: 13px;
    cursor: pointer;
    user-select: none;
    transition: all 0.12s ease;
  }
  .pill:hover { border-color: var(--gray-2); }
  .pill.active {
    background: var(--navy);
    color: #fff;
    border-color: var(--navy);
  }
  .pill .count {
    display: inline-block;
    margin-left: 6px;
    padding: 1px 6px;
    border-radius: 10px;
    font-size: 11px;
    background: var(--mute-2);
    color: var(--gray);
    font-weight: 600;
  }
  .pill.active .count {
    background: rgba(255,255,255,0.16);
    color: #fff;
  }

  .search {
    margin-left: auto;
    flex: 0 0 280px;
    max-width: 100%;
  }
  .search input {
    width: 100%;
    padding: 8px 12px;
    border: 1px solid var(--mute);
    border-radius: 8px;
    font-size: 13px;
    color: var(--ink);
    background: #fff;
    outline: none;
    font-family: inherit;
  }
  .search input:focus {
    border-color: var(--emerald);
    box-shadow: 0 0 0 3px rgba(5,150,105,0.12);
  }

  table {
    width: 100%;
    border-collapse: collapse;
    background: #fff;
    border: 1px solid var(--mute);
    border-radius: 10px;
    overflow: hidden;
    font-size: 13px;
  }
  thead th {
    background: var(--navy);
    color: #fff;
    text-align: left;
    padding: 10px 12px;
    font-weight: 600;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    cursor: pointer;
    user-select: none;
    white-space: nowrap;
  }
  thead th.sortable:hover { background: #132b52; }
  thead th .arrow { opacity: 0.45; margin-left: 4px; font-size: 10px; }
  thead th.sort-asc .arrow,
  thead th.sort-desc .arrow { opacity: 1; }

  tbody td {
    padding: 10px 12px;
    border-top: 1px solid var(--mute);
    vertical-align: top;
  }
  tbody tr:hover { background: var(--mute-2); }

  .id-cell { font-family: "SF Mono", "Menlo", monospace; color: var(--gray); font-size: 11px; }
  .status-badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 999px;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    white-space: nowrap;
  }
  .status-Open { background: var(--mute-2); color: var(--gray); }
  .status-In\ Progress { background: #FEF3C7; color: var(--amber); }
  .status-Done { background: #D1FAE5; color: var(--green); }
  .status-Dropped { background: var(--mute-2); color: var(--gray-2); font-style: italic; }

  .pri-P0 { color: var(--red); font-weight: 600; }
  .pri-P1 { color: var(--amber); font-weight: 500; }
  .pri-P2 { color: var(--gray); }

  .due { font-family: "SF Mono", "Menlo", monospace; font-size: 12px; white-space: nowrap; }
  .due.overdue { color: var(--red); font-weight: 600; }
  .due.soon { color: var(--amber); font-weight: 500; }

  .owner-badge {
    display: inline-block;
    padding: 2px 7px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 600;
    font-family: "SF Mono", "Menlo", monospace;
  }
  .owner-MH { background: #DBEAFE; color: #1E40AF; }
  .owner-CC { background: #EDE9FE; color: #5B21B6; }

  .title-cell { max-width: 520px; }
  .title-cell .title-text { color: var(--ink); }
  .title-cell .note {
    display: block;
    color: var(--gray);
    font-size: 12px;
    margin-top: 3px;
    font-style: italic;
  }

  .section-cell { color: var(--gray); font-size: 12px; white-space: nowrap; }

  .empty {
    text-align: center;
    padding: 40px 20px;
    color: var(--gray);
    font-style: italic;
  }

  footer.foot {
    margin-top: 32px;
    padding-top: 20px;
    border-top: 1px solid var(--mute);
    font-size: 12px;
    color: var(--gray);
    display: flex;
    justify-content: space-between;
    gap: 20px;
    flex-wrap: wrap;
  }
  footer.foot code {
    background: var(--mute-2);
    padding: 2px 6px;
    border-radius: 4px;
    font-size: 11px;
  }

  @media (max-width: 900px) {
    .cards { grid-template-columns: repeat(2, 1fr); }
    .radar { grid-template-columns: 1fr 1fr; }
    .search { margin-left: 0; flex: 1 1 100%; }
  }
</style>
</head>
<body>
<div class="wrap">

<header class="top">
  <div class="brand">
    <div class="wordmark">tenu</div>
    <div class="subtitle">task tracker · single source of truth</div>
  </div>
  <div class="meta">
    <div class="launch">Soft launch Mon __LAUNCH_DATE__</div>
    <div class="countdown-big" id="countdown">__DAYS_TO_LAUNCH__ days</div>
    <div>Regenerated <span id="gen-ts">__GENERATED__</span></div>
  </div>
</header>

<section class="cards">
  <div class="card">
    <div class="label">Open</div>
    <div class="value" id="c-open">__OPEN_TOTAL__</div>
    <div class="hint"><span id="c-ip">__IN_PROGRESS__</span> in progress</div>
  </div>
  <div class="card accent-emerald">
    <div class="label">Done</div>
    <div class="value" id="c-done">__DONE_TOTAL__</div>
    <div class="hint">closed with dated note</div>
  </div>
  <div class="card accent-amber">
    <div class="label">Due next 7 days</div>
    <div class="value" id="c-due7">__DUE_7__</div>
    <div class="hint">launch-gate burn-down</div>
  </div>
  <div class="card accent-red">
    <div class="label">P0 remaining</div>
    <div class="value" id="c-p0">__P0_REMAINING__</div>
    <div class="hint">blocks 11 May soft launch</div>
  </div>
</section>

<section class="radar">
  <div>
    <h3>Risk radar</h3>
    <div class="tagline">What should bother you today. Tap a stat to filter.</div>
  </div>
  <div class="stat" data-filter="overdue" style="cursor:pointer;">
    <div class="n bad" id="r-overdue">__OVERDUE__</div>
    <div class="lbl">Overdue</div>
  </div>
  <div class="stat" data-filter="p0-no-due" style="cursor:pointer;">
    <div class="n warn" id="r-p0-no-due">__P0_NO_DUE__</div>
    <div class="lbl">P0 without due date</div>
  </div>
  <div class="stat">
    <div class="n">
      <span id="r-mh">__MH_OPEN__</span>
      <span style="color:#C6CEDB;font-size:18px;">·</span>
      <span id="r-cc">__CC_OPEN__</span>
    </div>
    <div class="lbl">MH · CC open</div>
  </div>
</section>

<div class="controls">
  <div class="pills" id="filter-pills">
    <span class="pill active" data-filter="all">All <span class="count" id="cnt-all">0</span></span>
    <span class="pill" data-filter="mh">MH <span class="count" id="cnt-mh">0</span></span>
    <span class="pill" data-filter="cc">CC <span class="count" id="cnt-cc">0</span></span>
    <span class="pill" data-filter="p0">P0 <span class="count" id="cnt-p0">0</span></span>
    <span class="pill" data-filter="p1">P1 <span class="count" id="cnt-p1">0</span></span>
    <span class="pill" data-filter="p2">P2 <span class="count" id="cnt-p2">0</span></span>
    <span class="pill" data-filter="overdue">Overdue <span class="count" id="cnt-overdue">0</span></span>
    <span class="pill" data-filter="next-7">Next 7 days <span class="count" id="cnt-next-7">0</span></span>
    <span class="pill" data-filter="in-progress">In progress <span class="count" id="cnt-in-progress">0</span></span>
    <span class="pill" data-filter="done">Done <span class="count" id="cnt-done">0</span></span>
  </div>
  <div class="search">
    <input type="search" id="search" placeholder="Search title, note, section...">
  </div>
</div>

<table id="task-table">
  <thead>
    <tr>
      <th class="sortable" data-col="id">ID <span class="arrow">↕</span></th>
      <th class="sortable" data-col="due">Due <span class="arrow">↕</span></th>
      <th class="sortable" data-col="priority">P <span class="arrow">↕</span></th>
      <th class="sortable" data-col="owner">Own <span class="arrow">↕</span></th>
      <th class="sortable" data-col="status">Status <span class="arrow">↕</span></th>
      <th class="sortable" data-col="section">Section <span class="arrow">↕</span></th>
      <th class="sortable" data-col="title">Title & note <span class="arrow">↕</span></th>
    </tr>
  </thead>
  <tbody id="task-tbody"></tbody>
</table>

<footer class="foot">
  <div>
    Master: <code>TASKS.md</code> · mirror: <code>Tenu-Reste-a-Faire.xlsx</code> · protocol: <code>CLAUDE.md</code>
  </div>
  <div>
    Refresh: <code>python3 scripts/tracker-refresh.py</code>
  </div>
</footer>

</div>

<script id="task-data" type="application/json">__DATA_JSON__</script>

<script>
(function() {
  const raw = document.getElementById("task-data").textContent;
  const payload = JSON.parse(raw);
  const tasks = payload.tasks;
  const summary = payload.summary;

  const today = new Date(summary.today + "T00:00:00");
  const horizon = new Date(today); horizon.setDate(horizon.getDate() + 7);

  function dateCmp(a, b) {
    if (!a && !b) return 0;
    if (!a) return 1;
    if (!b) return -1;
    return a < b ? -1 : a > b ? 1 : 0;
  }

  function priorityRank(p) {
    if (p === "P0 Blocker") return 0;
    if (p === "P1 Important") return 1;
    if (p === "P2 Post-launch") return 2;
    return 3;
  }

  const tbody = document.getElementById("task-tbody");
  const pills = document.querySelectorAll(".pill");
  const searchInput = document.getElementById("search");

  let activeFilter = "all";
  let sortCol = "due";
  let sortDir = 1;

  function isOpenLike(t) { return t.status === "Open" || t.status === "In Progress"; }
  function isOverdue(t) { return isOpenLike(t) && t.due && new Date(t.due + "T00:00:00") < today; }
  function isNext7(t) { return isOpenLike(t) && t.due && new Date(t.due + "T00:00:00") <= horizon; }

  const filterFns = {
    "all": t => isOpenLike(t),
    "mh": t => isOpenLike(t) && t.owner === "MH",
    "cc": t => isOpenLike(t) && t.owner === "CC",
    "p0": t => isOpenLike(t) && t.priority === "P0 Blocker",
    "p1": t => isOpenLike(t) && t.priority === "P1 Important",
    "p2": t => isOpenLike(t) && t.priority === "P2 Post-launch",
    "overdue": t => isOverdue(t),
    "next-7": t => isNext7(t),
    "p0-no-due": t => isOpenLike(t) && t.priority === "P0 Blocker" && !t.due,
    "in-progress": t => t.status === "In Progress",
    "done": t => t.status === "Done",
  };

  function updatePillCounts() {
    Object.keys(filterFns).forEach(k => {
      const el = document.getElementById("cnt-" + k);
      if (!el) return;
      el.textContent = tasks.filter(filterFns[k]).length;
    });
  }

  function escapeHtml(s) {
    return String(s || "").replace(/[&<>"']/g, c =>
      ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])
    );
  }

  function dueClass(t) {
    if (!t.due) return "";
    if (isOverdue(t)) return "overdue";
    if (isNext7(t)) return "soon";
    return "";
  }

  function priShort(p) {
    if (p === "P0 Blocker") return "P0";
    if (p === "P1 Important") return "P1";
    if (p === "P2 Post-launch") return "P2";
    return "";
  }

  function priClass(p) {
    if (p === "P0 Blocker") return "pri-P0";
    if (p === "P1 Important") return "pri-P1";
    if (p === "P2 Post-launch") return "pri-P2";
    return "";
  }

  function statusClass(s) {
    return "status-" + s.replace(/ /g, "\\ ");
  }

  function render() {
    const q = searchInput.value.trim().toLowerCase();
    let rows = tasks.filter(filterFns[activeFilter] || filterFns.all);
    if (q) {
      rows = rows.filter(t =>
        (t.title || "").toLowerCase().includes(q) ||
        (t.note || "").toLowerCase().includes(q) ||
        (t.section || "").toLowerCase().includes(q) ||
        (t.id || "").toLowerCase().includes(q)
      );
    }

    rows.sort((a, b) => {
      let av, bv;
      if (sortCol === "priority") {
        av = priorityRank(a.priority);
        bv = priorityRank(b.priority);
      } else if (sortCol === "due") {
        av = a.due || "9999-99-99";
        bv = b.due || "9999-99-99";
      } else {
        av = (a[sortCol] || "").toString().toLowerCase();
        bv = (b[sortCol] || "").toString().toLowerCase();
      }
      return dateCmp(av, bv) * sortDir;
    });

    if (!rows.length) {
      tbody.innerHTML = '<tr><td colspan="7" class="empty">No tasks match this filter.</td></tr>';
      return;
    }

    tbody.innerHTML = rows.map(t => `
      <tr>
        <td class="id-cell">${escapeHtml(t.id)}</td>
        <td class="due ${dueClass(t)}">${escapeHtml(t.due || "—")}</td>
        <td class="${priClass(t.priority)}">${priShort(t.priority) || "—"}</td>
        <td><span class="owner-badge owner-${t.owner}">${escapeHtml(t.owner)}</span></td>
        <td><span class="status-badge status-${t.status.replace(/ /g, '_')}" style="${statusBadgeStyle(t.status)}">${escapeHtml(t.status)}</span></td>
        <td class="section-cell">${escapeHtml(t.section)}</td>
        <td class="title-cell">
          <span class="title-text">${escapeHtml(t.title)}</span>
          ${t.note ? `<span class="note">${escapeHtml(t.note)}</span>` : ""}
        </td>
      </tr>
    `).join("");

    // Update sort arrows
    document.querySelectorAll("thead th.sortable").forEach(th => {
      th.classList.remove("sort-asc", "sort-desc");
      const arrow = th.querySelector(".arrow");
      if (th.dataset.col === sortCol) {
        th.classList.add(sortDir > 0 ? "sort-asc" : "sort-desc");
        arrow.textContent = sortDir > 0 ? "↑" : "↓";
      } else {
        arrow.textContent = "↕";
      }
    });
  }

  function statusBadgeStyle(s) {
    // inline fallback because CSS attribute-selector escaping is finicky
    if (s === "Open") return "background:#F3F4F6;color:#6B7280";
    if (s === "In Progress") return "background:#FEF3C7;color:#B45309";
    if (s === "Done") return "background:#D1FAE5;color:#047857";
    if (s === "Dropped") return "background:#F3F4F6;color:#9CA3AF;font-style:italic";
    return "";
  }

  pills.forEach(p => p.addEventListener("click", () => {
    pills.forEach(x => x.classList.remove("active"));
    p.classList.add("active");
    activeFilter = p.dataset.filter;
    render();
  }));

  // Radar stats clickable
  document.querySelectorAll(".radar .stat[data-filter]").forEach(el => {
    el.addEventListener("click", () => {
      const f = el.dataset.filter;
      const pill = document.querySelector('.pill[data-filter="' + f + '"]');
      if (pill) {
        pill.click();
      } else {
        // No matching pill (p0-no-due has no pill) — activate synthetic filter
        pills.forEach(x => x.classList.remove("active"));
        activeFilter = f;
        render();
      }
    });
  });

  document.querySelectorAll("thead th.sortable").forEach(th => {
    th.addEventListener("click", () => {
      const col = th.dataset.col;
      if (sortCol === col) sortDir = -sortDir;
      else { sortCol = col; sortDir = 1; }
      render();
    });
  });

  searchInput.addEventListener("input", render);

  // Urgent countdown colouring
  if (summary.days_to_launch < 14) {
    document.getElementById("countdown").classList.add("urgent");
  }
  const cd = document.getElementById("countdown");
  cd.textContent = summary.days_to_launch + (summary.days_to_launch === 1 ? " day" : " days");

  updatePillCounts();
  render();
})();
</script>

</body>
</html>
"""


def main() -> int:
    if not TASKS_MD.exists():
        print(f"ERROR: {TASKS_MD} not found", file=sys.stderr)
        return 1
    text = TASKS_MD.read_text(encoding="utf-8")
    tasks, meta = parse_tasks(text)
    if not tasks:
        print("WARNING: no tasks parsed from TASKS.md", file=sys.stderr)
    render(tasks, meta, XLSX)
    render_html(tasks, meta, DASHBOARD_HTML)
    open_n = sum(1 for t in tasks if t.status in ("Open", "In Progress"))
    done_n = sum(1 for t in tasks if t.status == "Done")
    dropped_n = sum(1 for t in tasks if t.status == "Dropped")
    print(f"OK  {XLSX.name}  open={open_n}  done={done_n}  dropped={dropped_n}  total={len(tasks)}")
    print(f"OK  {DASHBOARD_HTML.name}  open file://{DASHBOARD_HTML}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
